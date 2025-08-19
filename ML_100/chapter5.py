import pandas as pd
import os
import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Font
from IPython.display import display, clear_output
from ipywidgets import DatePicker

m_store = pd.read_csv("d:\\data_analysis_lab\\ML_100\\data\\99_master\\m_store.csv")
m_area = pd.read_csv("d:\\data_analysis_lab\\ML_100\\data\\99_master\\m_area.csv")
tbl_order_4 = pd.read_csv("./ML_100/data/0_input/tbl_order_202104.csv")
tbl_order_5 = pd.read_csv("./ML_100/data/0_input/tbl_order_202105.csv")
tbl_order_6 = pd.read_csv("./ML_100/data/0_input/tbl_order_202106.csv")
tbl_order_7 = pd.read_csv("./ML_100/data/0_input/tbl_order_202107.csv")

#print(m_store.head())
#print(tbl_order_7.head())

# 기본 폴더 만들기 

data_dir = "d:\\data_analysis_lab\\ML_100\\data"
input_dir = os.path.join(data_dir, '0_input')
output_dir = os.path.join(data_dir, '10_output')
master_dir = os.path.join(data_dir, '99_master')

#print(input_dir)

# 폴더 만들기
os.makedirs(input_dir, exist_ok=True)
os.makedirs(output_dir, exist_ok=True)
os.makedirs(master_dir, exist_ok=True)

# 입력 데이터 확인 구조 
m_area_file = 'm_area.csv'
m_store_file = 'm_store.csv'
m_area = pd.read_csv(os.path.join(master_dir, m_area_file))
m_store = pd.read_csv(os.path.join(master_dir, m_store_file))

#m_area.head(3)

tg_ym = "202107"
target_file = 'tbl_order_' + tg_ym + ".csv"
target_data = pd.read_csv(os.path.join(input_dir, target_file))
#print(target_data.head())


max_date = pd.to_datetime(target_data['order_accept_date']).max()
min_date = pd.to_datetime(target_data['order_accept_date']).min()
max_str_date = max_date.strftime('%Y%m')
min_str_date = min_date.strftime('%Y%m')

if tg_ym == min_str_date and tg_ym == max_str_date :
    print("날짜가 일치합니다.")
else : 
    raise Exception("날짜가 일치하지 않습니다")


def calc_delta(t) : 
    t1, t2 = t
    delta = t2-t1
    return delta.total_seconds() / 60

def init_tran_df(trg_df) : 
    # 유지 보수용 매장 데이터 삭제 
    trg_df = trg_df.loc[trg_df['store_id'] != 999]
    trg_df = pd.merge(trg_df, m_store, on='store_id', how='left')
    trg_df = pd.merge(trg_df, m_area, on='area_cd', how='left')
    
    # 마스터 데이터에 없는 코드 대응 문자열 설정 
    trg_df.loc[trg_df['takeout_flag'] == 0, 'takeout_name'] = 'delivery'
    trg_df.loc[trg_df['takeout_flag'] == 1, 'takeout_name'] = 'takeout'
    
    trg_df.loc[trg_df['status'] == 0, 'status_name'] = '주문접수'
    trg_df.loc[trg_df['status'] == 1, 'status_name'] = '지불완료'
    trg_df.loc[trg_df['status'] == 2, 'status_name'] = '배달완료'
    trg_df.loc[trg_df['status'] == 9, 'status_name'] = '주문취소'
    
    trg_df.loc[:, 'order_date'] = pd.to_datetime(trg_df['order_accept_date']).dt.date
    
    # 배달시간 계산
    trg_df['order_accept_datetime'] = pd.to_datetime(trg_df['order_accept_date'])
    trg_df['delivered_datetime'] = pd.to_datetime(trg_df['delivered_date'])
    trg_df.loc[:,'delta'] = trg_df[['order_accept_datetime', 'delivered_datetime']].apply(calc_delta,axis=1)
    
    return trg_df

# 해당 월 데이터 초기화
target_data = init_tran_df(target_data)
print("============= 해당 월 데이터 초기화 ==============")
print(target_data.head())
print("============= 해당 월 데이터 초기화 ==============")

def get_rank_df(target_data) : 
    #매잘 데이터 작성, 순위 DF 변환
    tmp = target_data.loc[target_data['status'].isin([1,2])]
    rank = tmp.groupby(['store_id'])['total_amount'].sum().sort_values(ascending=False)
    rank = pd.merge(rank, m_store, on='store_id', how='left')
    return rank

def get_cancel_rank_df(target_data) : 
    # 주문 취소율 계산, 순위 DF 반환
    cancel_df = pd.DataFrame()
    cancel_cnt = target_data.loc[target_data['status'] == 9].groupby(['store_id'])['store_id'].count()
    order_cnt = target_data.loc[target_data['status'].isin([1,2,9])].groupby(['store_id'])['store_id'].count()
    cancel_rate = (cancel_cnt / order_cnt) * 100
    cancel_df['cancel_rate'] = cancel_rate
    cancel_df = pd.merge(cancel_df, m_store, on='store_id', how='left')
    cancel_df = cancel_df.sort_values('cancel_rate', ascending=True)
    return cancel_df

def data_export(df, ws, row_start, col_start) : 
    # 스타일 정의 
    side = Side(style='thin', color='008080')
    border = Border(top=side, bottom=side, left=side, right=side)
    
    rows = dataframe_to_rows(df, index=False, header=True)
    
    for row_no, row in enumerate(rows, row_start) : 
        for col_no, value in enumerate(row, col_start) :
            cell = ws.cell(row_no, col_no)
            cell.value = value
            cell.border = border
            if row_no == row_start : 
                cell.fill = PatternFill(patternType='solid', fgColor='008080')
                cell.font = Font(bold=True, color='FFFFFF')
    
# 본부용 보고서 데이터 처리 
def make_report_hq(target_data, output_folder) : 
    rank =  get_rank_df(target_data)
    cacel_rank = get_cancel_rank_df(target_data)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws.title = '요약보고서(본부용)'
    
    cell = ws.cell(1,1)
    cell.value = f'본부용 {max_str_date}월분 요약 보고서'
    cell.font = Font(bold=True, color='008080', size=20)
         
    cell = ws.cell(3,2)
    cell.value = f'{max_str_date}월분 매출 총액'
    cell.font = Font(bold=True, color='008080', size=20)
    
    cell = ws.cell(3,6)
    cell.value = f"{'{:,}'.format(rank['total_amount'].sum())}"
    cell.font = Font(bold=True, color='008080', size=20)
    
    # 매출 순뤼 직접 출력
    cell = ws.cell(5,2)
    cell.value = f'매출순위'
    cell.font = Font(bold=True, color='008080', size=15)
    
    data_export(rank, ws, 6, 2)
    
    cell = ws.cell(5,8)
    cell.value = f'주문 취소율 순위'
    cell.font = Font(bold=True, color='008080', size=16)
    
    data_export(cacel_rank, ws, 6, 8)
    
    wb.save(os.path.join(output_folder, f'export_hq_{max_str_date}.xlsx'))
    wb.close()
    
    
def get_store_rank(target_id, target_df) : 
    rank = get_rank_df(target_df)
    store_rank = rank.loc[rank['store_id'] == target_id].index + 1
    return store_rank[0]

def get_store_sale(target_id, target_df) : 
    rank = get_rank_df(target_df)
    store_sale = rank.loc[rank['store_id'] == target_id]['total_amount']
    return store_sale

def get_store_cancel_rank(target_id, target_df) :
    cancel_df = get_cancel_rank_df(target_df)
    cancel_df = cancel_df.reset_index()
    store_cancel_rank = cancel_df.loc[cancel_df['store_id'] == target_id].index + 1
    return store_cancel_rank[0]

def get_store_cancel_count(target_id, target_df) : 
    store_cancel_count = target_df.loc[(target_df['status'] == 9) &
                                       (target_df['store_id'] == target_id)
                                       ].groupby(['store_id'])['store_id'].count()
    return store_cancel_count

def get_delivery_rank_df(target_id, target_df) :
    delivery = target_df.loc[target_df['status'] == 2]
    delivery_rank = delivery.groupby(['store_id'])['delta'].mean().sort_values()
    delivery_rank = pd.merge(delivery_rank, m_store, on='store_id', how='left')
    return delivery_rank

def get_delivery_rank_store(target_id, target_df) : 
    delivery_rank = get_delivery_rank_df(target_id, target_df)
    store_delivery_rank = delivery_rank.loc[delivery_rank['store_id'] == target_id].index + 1
    return store_delivery_rank[0]

def make_report_hq_r2(target_date_list, output_folder) : 
    wb = openpyxl.Workbook()
    
    file_date = ''

    for tmp in target_date_list : 
        df = pd.DataFrame(tmp)
        
        df_date = pd.to_datetime(df['order_accept_date']).max()
        trg_date = df_date.strftime("%Y%m")
        
        if file_date ==  "":
            # 처음에만 파일허용
            file_date = trg_date
        
        rank = get_rank_df(df)
        cancel_rank = get_cancel_rank_df(df)
        
        ws = wb.create_sheet(title=f'{trg_date}월분')
        
        cell = ws.cell(1,1)
        cell.value = f'본부용 {trg_date}월분 요약 보고서'
        cell.font = Font(bold=True, color="008080", size=20)
        
        cell = ws.cell(3,2)
        cell.value = f'{max_str_date}월분 매출 총액'
        cell.font = Font(bold=True, color="008080", size=20)
        
        cell = ws.cell(3,6)
        cell.value = f'{'{:,}'.format(rank['total_amount'].sum())}'
        cell.font = Font(bold=True, color="008080", size=20)
        
        cell = ws.cell(5,2)
        cell.value = f'매출순위'
        cell.font = Font(bold=True, color='0008080', size=16)
        
        data_export(rank, ws, 6, 2)
        
        # 주문 취소율 순위를 직접 출력
        cell = ws.cell(5,8)
        cell.value = f'주문 취소율 순위'
        cell.font = Font(bold=True, color='008080', size=16)
        
        data_export(cancel_rank, ws, 6,8)
    

def make_report_store_r2(target_data, target_id, output_folder) : 
    rank = get_store_rank(target_id, target_data)
    sale = get_store_sale(target_id, target_data)
    cancel_rank = get_store_cancel_rank(target_id, target_data)
    cancel_count = get_store_cancel_count(target_id, target_data)
    delivery_df = get_delivery_rank_df(target_id, target_data)
    delivery_rank = get_delivery_rank_store(target_id, target_data)
    
    store_name = m_store.loc[m_store['store_id'] == target_id]['store_name'].values[0]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '매장용 보고서'
    
    cell = ws.cell(1,1)
    cell.value = f'{store_name} {max_str_date}월분 요약 보고서'
    cell.font = Font(bold=True, color='008080', size=20)
    
    cell = ws.cell(3,2)
    cell.value = f'{max_str_date}월분 매출 총액'
    cell.font = Font(bold=True, color='008080', size=20)
    
    cell = ws.cell(3,6)
    cell.value = f"{'{:,}'.format(sale.values[0])}"
    cell.font = Font(bold=True, color='008080', size=20)
    
    cell = ws.cell(5,2)
    cell.value = f'매출순위'
    cell.font = Font(bold=True, color='008080', size=15)
    
    cell = ws.cell(6,2)
    cell.value = f'{rank} 위'
    cell.font = Font(bold=True, color='008080', size=16)
    
    tmp_df = target_data.loc[(target_data['store_id'] == target_id)&
                             (target_data['status'].isin([1,2]))] 
    tmp_df = tmp_df[['order_accept_date', 'customer_id', 'total_amount', 'takeout_name', 'status_name']]
    data_export(tmp_df, ws, 7, 2)
    
    # 주문 취소율 순위 직접 출력        
    cell = ws.cell(5,8)
    cell.value = f'매출 취소율 순위'
    cell.font = Font(bold=True, color='008080', size=15)
    
    cell = ws.cell(5,12)
    cell.value = f'{cancel_rank}위, {cancel_count.values[0]}회'
    cell.font = Font(bold=True, color='008080', size=16)
    
    cell = ws.cell(6,8)
    cell.value = f'매출 취소 데이터'
    cell.font = Font(bold=True, color='008080', size=16)
    
    # 테이블 삽입
    tmp_df = target_data.loc[(target_data['store_id'] == target_id)&
                             (target_data['status'] == 9)]
     
    tmp_df = tmp_df[['order_accept_date', 'customer_id', 'total_amount', 'takeout_name', 'status_name']]
    data_export(tmp_df, ws, 7,8)
    
    # 배달 시간 직접 출력
    
    ave_time = delivery_df.loc[delivery_df['store_id'] == target_id]['delta'].values[0]
    cell = ws.cell(5,14)
    cell.value = f'배달완료 소요 시간 순위'
    cell.font = Font(bold=True, color="008080", size=16)

    cell = ws.cell(5,18)
    cell.value = f'{delivery_rank}위, 평균{ave_time}분'
    cell.font = Font(bold=True, color="008080", size=16)

    cell = ws.cell(6,14)
    cell.value = f'각 매장 배달 시간 순위'
    cell.font = Font(bold=True, color="008080", size=16)
    
    data_export(delivery_df, ws, 7, 14)
    
    wb.save(os.path.join(output_folder, f'{target_id}_{store_name}_report_{max_str_date}.xlsx'))
    wb.close()
    
    
# 본부용 보고서
make_report_hq(target_data, output_dir)

# 각 매장용 보고서
for store_id in m_store.loc[m_store['store_id'] != 999]['store_id'] :
    make_report_store_r2(target_data, store_id, output_dir)
      
      
def make_active_folder(targetYM) :
    now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    target_output_dir_name = targetYM + "_" + now
    target_output_dir = os.path.join(output_dir, target_output_dir_name)
    os.makedirs(target_output_dir)
    print(target_output_dir_name)
    return target_output_dir

target_output_dir = make_active_folder(tg_ym)

# 본부용 보고서(생성 위치 변경)
make_report_hq_r2(target_data, target_output_dir)

for store_id in m_store.loc[m_store['store_id'] != 999]['store_id'] :
    area_cd = m_store.loc[m_store['store_id'] == store_id]['area_cd']
    area_name = m_area.loc[m_area['area_cd'] == area_cd.values[0]]['narrow_area'].values[0]
    target_store_output_dir = os.path.join(target_output_dir, area_name)
    os.makedirs(target_store_output_dir, exist_ok=True)
    make_report_store(target_data, store_id, target_store_output_dir) 
    
def order_by_date(val) : 
    clear_output()
    display(date_picker)
    
    df_array = []
    
    print("데이터 확인. 데이터를 준비합니다")
    
    date_str = str(val['new'])
    date_dt = datetime.datetime.strptime(date_str, '%Y-%m-%d')
    target_ym = date_dt.strftime('%Y%m')
    
    # 폴더 자동 생성
    target_output_dir = make_active_folder(target_ym)
    
    # 선택한 기준 월 데이터 확인
    target_file = 'tbl_order_' + target_ym + '.csv'
    if os.path.exists(os.path.join(input_dir, target_file)) == False :
        print(f'{target_file}이 없습니다')
        return
    else : 
        df = pd.read_csv(os.path.join(input_dir, target_file))
        df = init_tran_df(df)
        df_array.append(df)
        
    target_ym_old = str(int(target_ym) - 1)
    target_file = 'tbl_order' + target_ym_old + ".csv"
    if os.path.exists(os.path.join(input_dir, target_file)) == True:
        df = pd.read_csv(os.path.join(input_dir, target_file))
        df = init_tran_df(df)
        df_array.append(df)
    
    print("데이터 분비 완료. 보고서를 생성합니다...")
    
    make_report_hq_r2()